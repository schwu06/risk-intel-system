"""Excel 报表导出：灰底表头、居中、列宽自适应。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from intl_ratings.models import REPORT_COLUMNS, IssuerRiskModel


def export_excel(
    rows: list[IssuerRiskModel],
    output_dir: Path,
    prefix: str = "国际评级监测",
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"{prefix}_{ts}.xlsx"

    records = [r.to_excel_dict() for r in rows]
    df = pd.DataFrame(records, columns=REPORT_COLUMNS)

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="国际评级")
        ws = writer.sheets["国际评级"]

        for col_idx in range(1, len(REPORT_COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

        for row_idx in range(2, len(records) + 2):
            for col_idx in range(1, len(REPORT_COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = center
                cell.border = thin

        for col_idx, col_name in enumerate(REPORT_COLUMNS, start=1):
            max_len = len(col_name)
            for row_idx in range(2, len(records) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, min(len(str(val)), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 48))

        ws.row_dimensions[1].height = 36
        ws.freeze_panes = "A2"

    return path
